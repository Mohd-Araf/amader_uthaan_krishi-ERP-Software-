# ==========================================================
# IMPORTS & CONSTANTS
# ==========================================================

from decimal import Decimal, ROUND_HALF_UP

from django.contrib import messages
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import user_passes_test, login_required
from django.core.paginator import Paginator
from django.db.models import Sum, Q
from django.utils import timezone

import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils import timezone
from django.http import HttpResponse
from apps.products.models import Order, Product
from apps.accounts.models import CustomUser

from .models import (
    Journal,
    JournalEntry,
    JournalItem,
    Account,
    ExpenseCategory,
    PaymentReceipt,
    Expense,
)

from .forms import (
    AccountForm,
    MakePaymentForm,
    VoucherCreateForm,
    JournalEntryFormSet,
)

from .services import (
    receive_customer_payment,
    make_supplier_payment,
    receive_customer_payment_voucher_wise,
    make_supplier_payment_voucher_wise,
    get_outstanding_vouchers,
    create_expense_voucher,
    create_contra_voucher,
    create_account_opening_journal,
)

TWO_DECIMAL = Decimal("0.01")
ZERO = Decimal("0.00")


def _q(value):
    """Safely converts and quantizes any amount to 2 decimal places."""
    if value is None:
        return ZERO
    return Decimal(str(value)).quantize(TWO_DECIMAL, rounding=ROUND_HALF_UP)

def _get_account_balance(account, entry_filter):
    """
    Calculates exact ledger balance for an account up to the specified entry filter,
    including its opening balance, rounded to 2 decimal places.
    Prevents double-counting if an Opening Balance Journal Voucher already exists.
    """
    # Check if an Opening Balance Journal Voucher exists for this account
    has_op_journal = JournalEntry.objects.filter(
        account=account,
        journal__status="posted",
        journal__notes__icontains="Opening Balance"
    ).exists()

    totals = JournalEntry.objects.filter(entry_filter, account=account).aggregate(
        dr=Sum("debit"),
        cr=Sum("credit")
    )
    t_dr = _q(totals["dr"])
    t_cr = _q(totals["cr"])

    if has_op_journal:
        op_dr = ZERO
        op_cr = ZERO
    else:
        op_dr = _q(account.opening_balance) if account.opening_balance_type == "debit" else ZERO
        op_cr = _q(account.opening_balance) if account.opening_balance_type == "credit" else ZERO

    net_dr = op_dr + t_dr
    net_cr = op_cr + t_cr

    if account.type1 in ["asset", "expense"]:
        bal = _q(net_dr - net_cr)
    else:
        bal = _q(net_cr - net_dr)

    # Custom attributes for template compatibility
    account.filtered_balance = bal
    account._filtered_balance = bal
    account.balance = bal
    return bal
# ==========================================================
# ADMIN CHECK PERMISSION
# ==========================================================

def admin_required(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


# ==========================================================
# FINANCIAL DASHBOARD
# ==========================================================

@user_passes_test(admin_required)
def financial_dashboard(request):
    """
    Overview of sales, discounts, charges, and account statistics.
    """
    sales_entries = JournalEntry.objects.filter(journal__voucher_type="sales", journal__status="posted")

    total_sale = _q(sales_entries.filter(account__type2="product").aggregate(total=Sum("credit"))["total"])
    total_discount = _q(sales_entries.filter(account__type2="discount").aggregate(total=Sum("debit"))["total"])
    total_packaging = _q(sales_entries.filter(account__type2="packaging").aggregate(total=Sum("credit"))["total"])
    total_bkash = _q(sales_entries.filter(account__type2="bkash_charge").aggregate(total=Sum("credit"))["total"])

    latest_journals = Journal.objects.select_related("created_by", "order").order_by("-created_at")[:25]

    context = {
        "total_sale": total_sale,
        "total_discount": total_discount,
        "total_packaging": total_packaging,
        "total_bkash": total_bkash,
        "total_orders": Order.objects.count(),
        "total_customers": CustomUser.objects.filter(is_staff=False, is_superuser=False).count(),
        "latest_journals": latest_journals,
    }

    return render(request, "finance/financial_dashboard.html", context)


# ==========================================================
# JOURNAL LIST & DETAIL
# ==========================================================

@user_passes_test(admin_required)
def journal_list(request):
    search = request.GET.get("search", "").strip()

    journals = Journal.objects.select_related("created_by", "order").order_by("-created_at")

    if search:
        journals = journals.filter(
            Q(journal_id__icontains=search)
            | Q(voucher_no__icontains=search)
            | Q(reference_no__icontains=search)
            | Q(order__order_code__icontains=search)
        )

    paginator = Paginator(journals, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "finance/journal_list.html",
        {
            "page_obj": page_obj,
            "search": search,
        },
    )


@user_passes_test(admin_required)
def journal_detail(request, journal_id):
    journal = get_object_or_404(
        Journal.objects.select_related("created_by", "posted_by", "order"),
        Q(journal_id=journal_id) | Q(voucher_no=journal_id)
    )

    entries = journal.entries.select_related("account").order_by("id")
    items = journal.items.select_related("product")

    subtotal = _q(items.aggregate(total=Sum("amount"))["total"])

    return render(
        request,
        "finance/journal_detail.html",
        {
            "journal": journal,
            "entries": entries,
            "items": items,
            "subtotal": subtotal,
            "total_debit": journal.total_debit,
            "total_credit": journal.total_credit,
        },
    )


# ==========================================================
# CHART OF ACCOUNTS (ACCOUNT CRUD)
# ==========================================================

@login_required
@user_passes_test(admin_required)
def account_list(request):
    accounts = Account.objects.all()

    search = request.GET.get("search", "").strip()
    type1 = request.GET.get("type1", "").strip()
    type2 = request.GET.get("type2", "").strip()
    status = request.GET.get("status", "").strip()

    if search:
        accounts = accounts.filter(
            Q(account_code__icontains=search)
            | Q(name__icontains=search)
            | Q(type1__icontains=search)
            | Q(type2__icontains=search)
        )

    if type1:
        accounts = accounts.filter(type1=type1)
    if type2:
        accounts = accounts.filter(type2=type2)
    if status:
        accounts = accounts.filter(status=status)

    accounts = accounts.order_by("type1", "type2", "name")

    return render(
        request,
        "finance/account_list.html",
        {
            "accounts": accounts,
            "search": search,
            "type1": type1,
            "type2": type2,
            "status": status,
        },
    )


@login_required
@user_passes_test(admin_required)
def account_create(request):
    if request.method == "POST":
        form = AccountForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    account = form.save()
                    create_account_opening_journal(
                        account=account,
                        created_by=request.user,
                    )
                messages.success(request, "Account created successfully.")
                return redirect("account_list")
            except Exception as e:
                messages.error(request, f"Error creating account: {e}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = AccountForm()

    return render(
        request,
        "finance/account_form.html",
        {"form": form, "title": "Create Account"},
    )


@login_required
@user_passes_test(admin_required)
def account_update(request, pk):
    account = get_object_or_404(Account, pk=pk)

    if request.method == "POST":
        form = AccountForm(request.POST, instance=account)
        if form.is_valid():
            form.save()
            messages.success(request, "Account updated successfully.")
            return redirect("account_list")
    else:
        form = AccountForm(instance=account)

    return render(
        request,
        "finance/account_form.html",
        {"form": form, "title": "Update Account"},
    )


@login_required
@user_passes_test(admin_required)
def account_delete(request, pk):
    account = get_object_or_404(Account, pk=pk)

    if request.method == "POST":
        account.delete()
        messages.success(request, "Account deleted successfully.")
        return redirect("account_list")

    return render(
        request,
        "finance/account_delete.html",
        {"account": account},
    )


@login_required
@user_passes_test(admin_required)
def account_view(request, pk):
    account = get_object_or_404(Account, pk=pk)
    return redirect("ledger_detail", account_id=account.id)


# ==========================================================
# VOUCHER-WISE RECEIVE PAYMENT
# ==========================================================

@user_passes_test(admin_required)
def receive_payment(request, account_id=None, journal_id=None):
    selected_customer = None
    target_voucher_id = None

    query_journal_id = journal_id or request.GET.get("journal_id")
    query_customer_id = account_id or request.GET.get("customer_id")

    if query_journal_id:
        target_journal = Journal.objects.filter(
            Q(journal_id=query_journal_id) | Q(voucher_no=query_journal_id)
        ).first()
        if target_journal:
            target_voucher_id = target_journal.id
            if target_journal.customer:
                selected_customer = target_journal.customer
            else:
                cust_entry = target_journal.entries.filter(account__type2="customer").first()
                if cust_entry:
                    selected_customer = cust_entry.account
    elif query_customer_id:
        selected_customer = Account.objects.filter(id=query_customer_id, type2="customer").first()

    customer_accounts = Account.objects.filter(type2="customer", status="active").order_by("name")
    cash_bank_accounts = Account.objects.filter(type2__in=["cash", "bank"], status="active").order_by("name")

    outstanding_vouchers = []
    if selected_customer:
        outstanding_vouchers = get_outstanding_vouchers(selected_customer)

        if target_voucher_id:
            outstanding_vouchers = [
                v for v in outstanding_vouchers
                if v.get("journal") and v["journal"].id == target_voucher_id
            ]

    if request.method == "POST":
        cust_id = request.POST.get("customer_account") or request.POST.get("customer")
        payment_acc_id = request.POST.get("payment_account")
        remarks = request.POST.get("remarks", "")

        try:
            customer_acc = get_object_or_404(Account, id=cust_id)
            payment_acc = Account.objects.filter(
                id=payment_acc_id).first() if payment_acc_id else Account.objects.filter(type2="cash",
                                                                                         is_default=True).first()

            voucher_ids = request.POST.getlist("voucher_ids")
            payment_lines = []

            for v_id in voucher_ids:
                amt_str = request.POST.get(f"amount_{v_id}", "0").strip()
                amt = _q(amt_str or "0")
                if amt > ZERO:
                    payment_lines.append({
                        "voucher_id": v_id,
                        "amount": amt,
                    })

            if not payment_lines:
                messages.error(request, "Please enter a valid amount greater than zero for at least one voucher.")
                redirect_url = request.path
                if query_journal_id:
                    redirect_url += f"?journal_id={query_journal_id}"
                elif customer_acc:
                    redirect_url += f"?customer_id={customer_acc.id}"
                return redirect(redirect_url)

            journal_obj = receive_customer_payment_voucher_wise(
                customer_account=customer_acc,
                payment_account=payment_acc,
                payment_lines=payment_lines,
                received_by=request.user,
                remarks=remarks,
            )

            messages.success(request, f"Payment received successfully. Voucher #{journal_obj.voucher_no} created.")
            return redirect("ledger_detail", account_id=customer_acc.id)

        except Exception as e:
            messages.error(request, f"Error processing payment: {e}")

    return render(
        request,
        "finance/payment_receive.html",
        {
            "selected_customer": selected_customer,
            "target_voucher_id": target_voucher_id,
            "customer_accounts": customer_accounts,
            "cash_bank_accounts": cash_bank_accounts,
            "outstanding_vouchers": outstanding_vouchers,
        },
    )


# ==========================================================
# VOUCHER-WISE MAKE PAYMENT
# ==========================================================

@user_passes_test(admin_required)
def make_payment_view(request, account_id=None, journal_id=None):
    selected_party = None
    target_voucher_id = None

    query_journal_id = journal_id or request.GET.get("journal_id")
    query_party_id = account_id or request.GET.get("party_id")

    if query_journal_id:
        target_journal = Journal.objects.filter(
            Q(journal_id=query_journal_id) | Q(voucher_no=query_journal_id)
        ).first()
        if target_journal:
            target_voucher_id = target_journal.id
            if target_journal.customer:
                selected_party = target_journal.customer
            else:
                party_entry = target_journal.entries.filter(
                    Q(account__type1__in=["liability", "expense"]) | Q(account__type2__in=["supplier", "payable"])
                ).first()
                if party_entry:
                    selected_party = party_entry.account
    elif query_party_id:
        selected_party = Account.objects.filter(id=query_party_id).first()

    party_accounts = Account.objects.filter(
        Q(type1__in=["liability", "expense"]) | Q(type2__in=["supplier", "payable"]),
        status="active"
    ).order_by("name")

    cash_bank_accounts = Account.objects.filter(type2__in=["cash", "bank"], status="active").order_by("name")

    outstanding_vouchers = []
    if selected_party:
        outstanding_vouchers = get_outstanding_vouchers(selected_party)

        if target_voucher_id:
            outstanding_vouchers = [
                v for v in outstanding_vouchers
                if v.get("journal") and v["journal"].id == target_voucher_id
            ]

    if request.method == "POST":
        p_acc_id = request.POST.get("party_account") or request.POST.get("account")
        payment_acc_id = request.POST.get("payment_account")
        remarks = request.POST.get("remarks", "")

        try:
            party_acc = get_object_or_404(Account, id=p_acc_id)
            payment_acc = Account.objects.filter(
                id=payment_acc_id).first() if payment_acc_id else Account.objects.filter(type2="cash",
                                                                                         is_default=True).first()

            voucher_ids = request.POST.getlist("voucher_ids")
            payment_lines = []

            for v_id in voucher_ids:
                amt_str = request.POST.get(f"amount_{v_id}", "0").strip()
                amt = _q(amt_str or "0")
                if amt > ZERO:
                    payment_lines.append({
                        "voucher_id": v_id,
                        "amount": amt,
                    })

            if not payment_lines:
                messages.error(request, "Please enter a valid amount greater than zero for at least one voucher.")
                redirect_url = request.path
                if query_journal_id:
                    redirect_url += f"?journal_id={query_journal_id}"
                elif party_acc:
                    redirect_url += f"?party_id={party_acc.id}"
                return redirect(redirect_url)

            journal_obj = make_supplier_payment_voucher_wise(
                party_account=party_acc,
                payment_account=payment_acc,
                payment_lines=payment_lines,
                paid_by=request.user,
                remarks=remarks,
            )

            messages.success(request,
                             f"Payment of ৳{journal_obj.total_debit} recorded successfully. Voucher #{journal_obj.voucher_no} created.")
            return redirect("ledger_detail", account_id=party_acc.id)

        except Exception as e:
            messages.error(request, f"Error recording payment: {e}")

    return render(
        request,
        "finance/make_payment.html",
        {
            "selected_party": selected_party,
            "target_voucher_id": target_voucher_id,
            "party_accounts": party_accounts,
            "cash_bank_accounts": cash_bank_accounts,
            "outstanding_vouchers": outstanding_vouchers,
        },
    )


# ==========================================================
# OUTSTANDING VOUCHERS API
# ==========================================================

@user_passes_test(admin_required)
def outstanding_vouchers_api(request, account_id):
    account = get_object_or_404(Account, id=account_id)
    vouchers = get_outstanding_vouchers(account)

    data = [
        {
            "voucher_id": item["journal"].id if item.get("journal") else "OPENING-BALANCE",
            "voucher_no": item.get("voucher_no") or (
                item["journal"].voucher_no if item.get("journal") else "OPENING-BALANCE"),
            "voucher_type": item.get("voucher_type") or (
                item["journal"].get_voucher_type_display() if item.get("journal") else "Opening Balance"),
            "date": item["journal"].created_at.strftime("%d-%m-%Y") if item.get("journal") else "Opening Balance",
            "original_amount": str(item["original_amount"]),
            "paid_amount": str(item["paid_amount"]),
            "outstanding_amount": str(item["outstanding_amount"]),
            "is_opening": item.get("is_opening", False),
        }
        for item in vouchers
    ]
    return JsonResponse(data, safe=False)


# ==========================================================
# CREATE CONTRA VOUCHER
# ==========================================================

@user_passes_test(admin_required)
def create_contra(request):
    cash_bank_accounts = Account.objects.filter(type2__in=["cash", "bank"], status="active").order_by("name")

    if request.method == "POST":
        from_id = request.POST.get("from_account")
        to_id = request.POST.get("to_account")
        amount = request.POST.get("amount")
        remarks = request.POST.get("remarks", "")

        try:
            from_acc = get_object_or_404(Account, id=from_id)
            to_acc = get_object_or_404(Account, id=to_id)

            journal = create_contra_voucher(
                from_account=from_acc,
                to_account=to_acc,
                amount=_q(amount),
                user=request.user,
                remarks=remarks,
            )

            messages.success(request, f"Contra Transfer Voucher {journal.voucher_no} created successfully.")
            return redirect("voucher_detail", voucher_no=journal.voucher_no)

        except Exception as e:
            messages.error(request, f"Contra error: {e}")

    return render(
        request,
        "finance/contra_form.html",
        {"accounts": cash_bank_accounts},
    )

from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Sum, Q
from django.db.models.functions import Coalesce
# finance অ্যাপের নিজস্ব models.py থেকে ইম্পোর্ট করুন
from .models import Account, JournalEntry, _q, ZERO
# আপনার প্রোজেক্টের ডেকোরেটর (যদি আলাদা অ্যাপে থাকে)
# from apps.accounts.decorators import admin_required


def _get_ledger_data(accounts_qs):
    """
    হেল্পার ফাংশন: N+1 ডাটাবেজ কুয়েরি প্রতিরোধ করে সব অ্যাকাউন্টের
    রিয়েল-টাইম ক্লোজিং ব্যালেন্স দ্রুত ক্যালকুলেট করার জন্য।
    """
    accounts = accounts_qs.annotate(
        total_dr=Coalesce(Sum('entries__debit', filter=Q(entries__journal__status='posted')), ZERO),
        total_cr=Coalesce(Sum('entries__credit', filter=Q(entries__journal__status='posted')), ZERO),
    )

    data = []
    for acc in accounts:
        # Opening Balance Journal আগে থেকেই আছে কিনা চেক
        has_op_journal = acc.entries.filter(
            journal__status="posted",
            journal__notes__icontains="Opening Balance"
        ).exists()

        if has_op_journal:
            op_debit = ZERO
            op_credit = ZERO
        else:
            op_debit = _q(acc.opening_balance) if acc.opening_balance_type == "debit" else ZERO
            op_credit = _q(acc.opening_balance) if acc.opening_balance_type == "credit" else ZERO

        net_debit = op_debit + _q(acc.total_dr)
        net_credit = op_credit + _q(acc.total_cr)

        if acc.type1 in ["asset", "expense"]:
            balance = _q(net_debit - net_credit)
        else:
            balance = _q(net_credit - net_debit)

        data.append({
            "account": acc,
            "balance": balance,
        })
    return data


# ==========================================================
# LEDGERS & REPORTS
# ==========================================================

@user_passes_test(admin_required)
def ledger_detail(request, account_id):
    account = get_object_or_404(Account, id=account_id)

    entries = JournalEntry.objects.filter(
        account=account,
        journal__status="posted"
    ).select_related("journal", "account").order_by("entry_date", "id")

    totals = entries.aggregate(total_debit=Sum("debit"), total_credit=Sum("credit"))
    total_debit = _q(totals["total_debit"])
    total_credit = _q(totals["total_credit"])

    has_op_journal = entries.filter(
        journal__notes__icontains="Opening Balance"
    ).exists()

    if has_op_journal:
        if account.type1 in ["asset", "expense"]:
            balance = _q(total_debit - total_credit)
        else:
            balance = _q(total_credit - total_debit)
    else:
        op_debit = _q(account.opening_balance) if account.opening_balance_type == "debit" else ZERO
        op_credit = _q(account.opening_balance) if account.opening_balance_type == "credit" else ZERO
        net_dr = op_debit + total_debit
        net_cr = op_credit + total_credit
        if account.type1 in ["asset", "expense"]:
            balance = _q(net_dr - net_cr)
        else:
            balance = _q(net_cr - net_dr)

    # Running Balance ক্যালকুলেশন
    running_balance = ZERO
    if not has_op_journal:
        if account.type1 in ["asset", "expense"]:
            running_balance = _q(account.opening_balance) if account.opening_balance_type == "debit" else _q(-account.opening_balance)
        else:
            running_balance = _q(account.opening_balance) if account.opening_balance_type == "credit" else _q(-account.opening_balance)

    for entry in entries:
        if account.type1 in ["asset", "expense"]:
            running_balance = _q(running_balance + entry.debit - entry.credit)
        else:
            running_balance = _q(running_balance + entry.credit - entry.debit)

        if running_balance > ZERO:
            if account.type1 in ["asset", "expense"]:
                entry.closing_dr = running_balance
                entry.closing_cr = ZERO
            else:
                entry.closing_cr = running_balance
                entry.closing_dr = ZERO
        elif running_balance < ZERO:
            abs_bal = _q(abs(running_balance))
            if account.type1 in ["asset", "expense"]:
                entry.closing_cr = abs_bal
                entry.closing_dr = ZERO
            else:
                entry.closing_dr = abs_bal
                entry.closing_cr = ZERO
        else:
            entry.closing_dr = ZERO
            entry.closing_cr = ZERO

    show_receive_payment = account.type2 in ["customer", "receivable"]
    show_add_payment = account.type2 in ["supplier", "payable"]
    is_debit_account = account.type1 in ["asset", "expense"]

    return render(
        request,
        "finance/ledger_detail.html",
        {
            "account": account,
            "entries": entries,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "balance": balance,
            "is_debit_account": is_debit_account,
            "show_receive_payment": show_receive_payment,
            "show_add_payment": show_add_payment,
        },
    )


@user_passes_test(admin_required)
def cash_ledger(request):

    cash_acc = Account.objects.filter(type2="cash", is_default=True).first()
    if not cash_acc:
        cash_acc = Account.objects.filter(type2="cash").first()
    if not cash_acc:
        cash_acc, _ = Account.objects.get_or_create(
            name="Cash Account",
            type1="asset",
            type2="cash",
            is_default=True
        )

    # ২. ক্যাশ অ্যাকাউন্টের সকল পোস্ট করা জার্নাল এন্ট্রি ফেচ করা
    entries = JournalEntry.objects.filter(
        account=cash_acc,
        journal__status="posted"
    ).select_related("journal", "account").order_by("entry_date", "id")

    # ৩. মোট ডেবিট (জমা) এবং ক্রেডিট (খরচ) সামারি
    totals = entries.aggregate(total_debit=Sum("debit"), total_credit=Sum("credit"))
    total_debit = _q(totals["total_debit"])
    total_credit = _q(totals["total_credit"])

    # ৪. ওপেনিং ব্যালেন্স চেক
    has_op_journal = entries.filter(
        journal__notes__icontains="Opening Balance"
    ).exists()

    if has_op_journal:
        op_debit = ZERO
        op_credit = ZERO
    else:
        op_debit = _q(cash_acc.opening_balance) if cash_acc.opening_balance_type == "debit" else ZERO
        op_credit = _q(cash_acc.opening_balance) if cash_acc.opening_balance_type == "credit" else ZERO

    # মোট ক্লোজিং ক্যাশ ব্যালেন্স (Debit - Credit)
    balance = _q((op_debit + total_debit) - (op_credit + total_credit))

    # ৫. প্রতিটি Transaction Line এর জন্য entry.running_balance ক্যালকুলেট করা
    running_balance = op_debit - op_credit
    for entry in entries:
        running_balance = _q(running_balance + entry.debit - entry.credit)
        entry.running_balance = running_balance  # cash_ledger.html টেমপ্লেটের জন্য সেট করা হলো

    # ৬. cash_ledger.html টেমপ্লেটে ডাটা রেন্ডার করা
    return render(
        request,
        "finance/cash_ledger.html",
        {
            "account": cash_acc,
            "entries": entries,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "balance": balance,
        },
    )

@user_passes_test(admin_required)
def bank_ledger(request):
    banks = Account.objects.filter(type2="bank").order_by("name")
    data = _get_ledger_data(banks)
    return render(
        request,
        "finance/cash_ledger.html",
        {"title": "Bank Ledger", "banks": data, "cashes": data, "accounts": data, "data": data}
    )


# ----------------------------------------------------------
# 2. PARTY / CUSTOMER / SUPPLIER / RECEIVABLE / PAYABLE -> finance/customer_ledger.html
# ----------------------------------------------------------

@user_passes_test(admin_required)
def customer_ledger(request):
    customers = Account.objects.filter(type2="customer").order_by("name")
    data = _get_ledger_data(customers)
    return render(
        request,
        "finance/customer_ledger.html",
        {"title": "Customer Ledger", "customers": data, "accounts": data, "data": data}
    )

@user_passes_test(admin_required)
def supplier_ledger(request):
    suppliers = Account.objects.filter(type2="supplier").order_by("name")
    data = _get_ledger_data(suppliers)
    return render(
        request,
        "finance/customer_ledger.html",
        {"title": "Supplier Ledger", "suppliers": data, "customers": data, "accounts": data, "data": data}
    )

@user_passes_test(admin_required)
def receivable_ledger(request):
    receivables = Account.objects.filter(type1="asset", type2="receivable").order_by("name")
    data = _get_ledger_data(receivables)
    return render(
        request,
        "finance/customer_ledger.html",
        {"title": "Receivable Ledger", "receivables": data, "customers": data, "accounts": data, "data": data}
    )

@user_passes_test(admin_required)
def payable_ledger(request):
    payables = Account.objects.filter(type1="liability", type2="payable").order_by("name")
    data = _get_ledger_data(payables)
    return render(
        request,
        "finance/customer_ledger.html",
        {"title": "Payable Ledger", "payables": data, "customers": data, "accounts": data, "data": data}
    )


# ----------------------------------------------------------
# 3. PRODUCT LEDGER -> finance/product_ledger.html
# ----------------------------------------------------------

@user_passes_test(admin_required)
def product_ledger(request):
    products = Account.objects.filter(type2="product").order_by("name")
    data = _get_ledger_data(products)
    return render(
        request,
        "finance/product_ledger.html",
        {"title": "Product Ledger", "products": data, "accounts": data, "data": data}
    )


# ----------------------------------------------------------
# 4. GENERAL & OTHER LEDGERS -> finance/other_ledger.html
# ----------------------------------------------------------

@user_passes_test(admin_required)
def asset_ledger(request):
    assets = Account.objects.filter(type1="asset").order_by("name")
    data = _get_ledger_data(assets)
    return render(
        request,
        "finance/other_ledger.html",
        {"title": "Asset Ledger", "assets": data, "others": data, "accounts": data, "data": data}
    )

@user_passes_test(admin_required)
def liability_ledger(request):
    liabilities = Account.objects.filter(type1="liability").order_by("name")
    data = _get_ledger_data(liabilities)
    return render(
        request,
        "finance/other_ledger.html",
        {"title": "Liability Ledger", "liabilities": data, "others": data, "accounts": data, "data": data}
    )

@user_passes_test(admin_required)
def equity_ledger(request):
    equities = Account.objects.filter(type1="equity").order_by("name")
    data = _get_ledger_data(equities)
    return render(
        request,
        "finance/other_ledger.html",
        {"title": "Equity Ledger", "equities": data, "others": data, "accounts": data, "data": data}
    )

@user_passes_test(admin_required)
def income_ledger(request):
    incomes = Account.objects.filter(type1="revenue").order_by("name")
    data = _get_ledger_data(incomes)
    return render(
        request,
        "finance/other_ledger.html",
        {"title": "Income Ledger", "incomes": data, "others": data, "accounts": data, "data": data}
    )

@user_passes_test(admin_required)
def expense_ledger(request):
    expenses = Account.objects.filter(type1="expense").order_by("name")
    data = _get_ledger_data(expenses)
    return render(
        request,
        "finance/other_ledger.html",
        {"title": "Expense Ledger", "expenses": data, "others": data, "accounts": data, "data": data}
    )

@user_passes_test(admin_required)
def other_ledger(request):
    others = Account.objects.exclude(
        type2__in=["customer", "product", "cash"]
    ).order_by("name")
    data = _get_ledger_data(others)
    return render(
        request,
        "finance/other_ledger.html",
        {"title": "Other Ledger", "others": data, "accounts": data, "data": data}
    )
# ==========================================================
# TRIAL BALANCE REPORT
# ==========================================================

@login_required
@user_passes_test(admin_required)
def trial_balance(request):
    search = request.GET.get("search", "").strip()
    type1 = request.GET.get("type", "").strip()
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()

    accounts = Account.objects.all()

    if search:
        accounts = accounts.filter(Q(account_code__icontains=search) | Q(name__icontains=search))
    if type1:
        accounts = accounts.filter(type1=type1)

    accounts = accounts.order_by("account_code")

    rows = []
    total_debit = ZERO
    total_credit = ZERO

    for account in accounts:
        entries = JournalEntry.objects.filter(account=account, journal__status="posted")
        if from_date:
            entries = entries.filter(entry_date__date__gte=from_date)
        if to_date:
            entries = entries.filter(entry_date__date__lte=to_date)

        totals = entries.aggregate(debit=Sum("debit"), credit=Sum("credit"))
        debit = _q(totals["debit"])
        credit = _q(totals["credit"])

        display_debit = ZERO
        display_credit = ZERO

        if account.type1 in ["asset", "expense"]:
            balance = debit - credit
            if balance >= ZERO:
                display_debit = balance
            else:
                display_credit = abs(balance)
        else:
            balance = credit - debit
            if balance >= ZERO:
                display_credit = balance
            else:
                display_debit = abs(balance)

        rows.append({
            "account": account,
            "debit": _q(display_debit),
            "credit": _q(display_credit),
        })

        total_debit = _q(total_debit + display_debit)
        total_credit = _q(total_credit + display_credit)

    return render(
        request,
        "finance/trial_balance.html",
        {
            "rows": rows,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "balanced": total_debit == total_credit,
            "search": search,
            "selected_type": type1,
            "from_date": from_date,
            "to_date": to_date,
        },
    )


# ==========================================================
# 🔥 NEW: INCOME STATEMENT REPORT
# ==========================================================

@login_required
@user_passes_test(admin_required)
def income_statement(request):
    """
    Income Statement (Profit & Loss Statement)
    Revenue - Expenses = Net Profit / Loss
    """
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()

    entries_q = Q(journal__status="posted")
    if from_date:
        entries_q &= Q(entry_date__date__gte=from_date)
    if to_date:
        entries_q &= Q(entry_date__date__lte=to_date)

    entries = JournalEntry.objects.filter(entries_q)

    # 1. REVENUE BREAKDOWN
    revenue_entries = entries.filter(account__type1="revenue")

    sales_rev = revenue_entries.filter(account__type2="product").aggregate(cr=Sum("credit"), dr=Sum("debit"))
    sales_revenue = _q((sales_rev["cr"] or ZERO) - (sales_rev["dr"] or ZERO))

    pkg_rev = revenue_entries.filter(account__type2="packaging").aggregate(cr=Sum("credit"), dr=Sum("debit"))
    packaging_revenue = _q((pkg_rev["cr"] or ZERO) - (pkg_rev["dr"] or ZERO))

    bkash_rev = revenue_entries.filter(account__type2="bkash_charge").aggregate(cr=Sum("credit"), dr=Sum("debit"))
    bkash_revenue = _q((bkash_rev["cr"] or ZERO) - (bkash_rev["dr"] or ZERO))

    other_rev = revenue_entries.exclude(account__type2__in=["product", "packaging", "bkash_charge"]).aggregate(
        cr=Sum("credit"), dr=Sum("debit"))
    other_revenue = _q((other_rev["cr"] or ZERO) - (other_rev["dr"] or ZERO))

    total_revenue = _q(sales_revenue + packaging_revenue + bkash_revenue + other_revenue)

    # 2. EXPENSES BREAKDOWN
    expense_entries = entries.filter(account__type1="expense")

    disc_exp = expense_entries.filter(account__type2="discount").aggregate(dr=Sum("debit"), cr=Sum("credit"))
    sales_discount = _q((disc_exp["dr"] or ZERO) - (disc_exp["cr"] or ZERO))

    operating_accounts = Account.objects.filter(type1="expense", status="active").exclude(type2="discount")
    operating_expenses_list = []
    total_operating_expense = ZERO

    for acc in operating_accounts:
        acc_entries = expense_entries.filter(account=acc).aggregate(dr=Sum("debit"), cr=Sum("credit"))
        amt = _q((acc_entries["dr"] or ZERO) - (acc_entries["cr"] or ZERO))
        if amt != ZERO:
            acc.filtered_balance = amt
            acc._filtered_balance = amt
            operating_expenses_list.append({
                "account": acc,
                "amount": amt,
                "name": acc.name,
                "total": amt,
            })
            total_operating_expense = _q(total_operating_expense + amt)

    total_expenses = _q(sales_discount + total_operating_expense)

    # 3. NET PROFIT / LOSS
    net_profit = _q(total_revenue - total_expenses)

    context = {
        "sales_revenue": sales_revenue,
        "total_sales": sales_revenue,
        "packaging_revenue": packaging_revenue,
        "total_packaging": packaging_revenue,
        "bkash_revenue": bkash_revenue,
        "total_bkash": bkash_revenue,
        "other_revenue": other_revenue,
        "total_revenue": total_revenue,

        "sales_discount": sales_discount,
        "total_discount": sales_discount,
        "operating_expenses": operating_expenses_list,
        "expense_breakdown": operating_expenses_list,
        "total_operating_expense": total_operating_expense,
        "total_expenses": total_expenses,
        "total_expense": total_expenses,

        "net_profit": net_profit,
        "is_profit": net_profit >= ZERO,
        "from_date": from_date,
        "to_date": to_date,
    }

    return render(request, "finance/income_statement.html", context)


# ==========================================================
# 🔥 NEW: BALANCE SHEET REPORT
# ==========================================================

@login_required
@user_passes_test(admin_required)
def balance_sheet(request):
    """
    Balance Sheet Statement
    Total Assets == Total Liabilities + Total Equity (Opening Equity + Current Profit)
    """
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()

    # Balance Sheet uses cumulative entries up to to_date (or all-time if no to_date)
    bs_entry_q = Q(journal__status="posted")
    if to_date:
        bs_entry_q &= Q(entry_date__date__lte=to_date)

    # Income Statement for Equity calculation uses the exact same date scope
    pnl_entry_q = Q(journal__status="posted")
    if from_date:
        pnl_entry_q &= Q(entry_date__date__gte=from_date)
    if to_date:
        pnl_entry_q &= Q(entry_date__date__lte=to_date)

    # Calculate Net Profit for Balance Sheet
    entries_pnl = JournalEntry.objects.filter(pnl_entry_q)
    rev_totals = entries_pnl.filter(account__type1="revenue").aggregate(cr=Sum("credit"), dr=Sum("debit"))
    total_rev = _q((rev_totals["cr"] or ZERO) - (rev_totals["dr"] or ZERO))

    exp_totals = entries_pnl.filter(account__type1="expense").aggregate(dr=Sum("debit"), cr=Sum("credit"))
    total_exp = _q((exp_totals["dr"] or ZERO) - (exp_totals["cr"] or ZERO))

    current_net_profit = _q(total_rev - total_exp)

    # 1. ASSETS
    asset_accounts = Account.objects.filter(type1="asset", status="active")

    cash_assets = [acc for acc in asset_accounts if acc.type2 == "cash"]
    bank_assets = [acc for acc in asset_accounts if acc.type2 == "bank"]
    customer_assets = [acc for acc in asset_accounts if acc.type2 in ["customer", "receivable"]]
    other_assets = [acc for acc in asset_accounts if acc.type2 not in ["cash", "bank", "customer", "receivable"]]

    total_cash = _q(sum(_get_account_balance(acc, bs_entry_q) for acc in cash_assets))
    total_bank = _q(sum(_get_account_balance(acc, bs_entry_q) for acc in bank_assets))
    total_customer = _q(sum(_get_account_balance(acc, bs_entry_q) for acc in customer_assets))
    total_other_assets = _q(sum(_get_account_balance(acc, bs_entry_q) for acc in other_assets))

    total_assets = _q(total_cash + total_bank + total_customer + total_other_assets)

    # 2. LIABILITIES
    liability_accounts = Account.objects.filter(type1="liability", status="active")

    supplier_liabilities = [acc for acc in liability_accounts if acc.type2 in ["supplier", "payable"]]
    other_liabilities = [acc for acc in liability_accounts if acc.type2 not in ["supplier", "payable"]]

    total_supplier = _q(sum(_get_account_balance(acc, bs_entry_q) for acc in supplier_liabilities))
    total_other_liabilities = _q(sum(_get_account_balance(acc, bs_entry_q) for acc in other_liabilities))

    total_liabilities = _q(total_supplier + total_other_liabilities)

    # 3. EQUITY
    equity_accounts = list(Account.objects.filter(type1="equity", status="active"))
    opening_equity = _q(sum(_get_account_balance(acc, bs_entry_q) for acc in equity_accounts))

    total_equity = _q(opening_equity + current_net_profit)

    # 4. TOTAL LIABILITIES & EQUITY
    total_liabilities_equity = _q(total_liabilities + total_equity)

    # 5. SYSTEM BALANCE CHECK
    difference = _q(total_assets - total_liabilities_equity)
    is_balanced = (difference == ZERO)

    context = {
        "cash_assets": cash_assets,
        "total_cash": total_cash,
        "bank_assets": bank_assets,
        "total_bank": total_bank,
        "customer_assets": customer_assets,
        "total_customer": total_customer,
        "other_assets": other_assets,
        "total_other_assets": total_other_assets,
        "total_assets": total_assets,

        "supplier_liabilities": supplier_liabilities,
        "total_supplier": total_supplier,
        "other_liabilities": other_liabilities,
        "total_other_liabilities": total_other_liabilities,
        "total_liabilities": total_liabilities,

        "equity_accounts": equity_accounts,
        "opening_equity": opening_equity,
        "current_net_profit": current_net_profit,
        "total_equity": total_equity,

        "total_liabilities_equity": total_liabilities_equity,
        "is_balanced": is_balanced,
        "difference": difference,
        "from_date": from_date,
        "to_date": to_date,
        "as_of_date": to_date,
    }

    return render(request, "finance/balance_sheet.html", context)


# ==========================================================
# MANUAL VOUCHER CREATION / EDIT / POST / CANCEL / DELETE
# ==========================================================

@user_passes_test(admin_required)
def voucher_list(request):
    vouchers = Journal.objects.select_related("created_by", "posted_by", "order").order_by("-created_at")

    search = request.GET.get("search", "").strip()
    voucher_type = request.GET.get("type", "").strip()
    status = request.GET.get("status", "").strip()

    if search:
        vouchers = vouchers.filter(
            Q(voucher_no__icontains=search)
            | Q(journal_id__icontains=search)
            | Q(reference_no__icontains=search)
        )
    if voucher_type:
        vouchers = vouchers.filter(voucher_type=voucher_type)
    if status:
        vouchers = vouchers.filter(status=status)

    paginator = Paginator(vouchers, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "finance/voucher_list.html",
        {
            "page_obj": page_obj,
            "search": search,
            "voucher_type": voucher_type,
            "status": status,
        },
    )


@user_passes_test(admin_required)
def voucher_detail(request, voucher_no):
    voucher = get_object_or_404(Journal, Q(voucher_no=voucher_no) | Q(journal_id=voucher_no))
    entries = voucher.entries.select_related("account").order_by("id")
    items = voucher.items.select_related("product")

    return render(
        request,
        "finance/voucher_detail.html",
        {
            "voucher": voucher,
            "journal": voucher,
            "entries": entries,
            "items": items,
            "total_debit": voucher.total_debit,
            "total_credit": voucher.total_credit,
        },
    )


@user_passes_test(admin_required)
@transaction.atomic
def voucher_create(request):
    if request.method == "POST":
        form = VoucherCreateForm(request.POST)
        if form.is_valid():
            voucher = form.save(commit=False)
            voucher.created_by = request.user
            voucher.status = "draft"
            voucher.save()

            formset = JournalEntryFormSet(request.POST, instance=voucher)
            if formset.is_valid():
                entries = formset.save(commit=False)
                if not entries:
                    messages.error(request, "At least one journal entry is required.")
                    voucher.delete()
                    return redirect("voucher_create")

                total_dr = sum(_q(e.debit) for e in entries)
                total_cr = sum(_q(e.credit) for e in entries)

                if total_dr != total_cr or total_dr <= ZERO:
                    messages.error(request, "Journal entries must be balanced (Debit == Credit > 0).")
                    voucher.delete()
                    return redirect("voucher_create")

                for entry in entries:
                    entry.journal = voucher
                    entry.created_by = request.user
                    entry.save()

                messages.success(request, f"Voucher {voucher.voucher_no} created successfully as draft.")
                return redirect("voucher_detail", voucher_no=voucher.voucher_no)
            voucher.delete()
    else:
        form = VoucherCreateForm()
        formset = JournalEntryFormSet()

    return render(
        request,
        "finance/voucher_form.html",
        {"form": form, "formset": formset, "title": "Create Manual Voucher"},
    )


@user_passes_test(admin_required)
@transaction.atomic
def voucher_edit(request, voucher_no):
    voucher = get_object_or_404(Journal, Q(voucher_no=voucher_no) | Q(journal_id=voucher_no))

    if voucher.status in ["posted", "cancelled"]:
        messages.error(request, "Posted or cancelled voucher cannot be edited.")
        return redirect("voucher_detail", voucher_no=voucher.voucher_no)

    if request.method == "POST":
        form = VoucherCreateForm(request.POST, instance=voucher)
        formset = JournalEntryFormSet(request.POST, instance=voucher)

        if form.is_valid() and formset.is_valid():
            voucher = form.save(commit=False)
            voucher.updated_at = timezone.now()
            voucher.save()

            entries = formset.save(commit=False)
            if not entries:
                messages.error(request, "At least two journal entries are required.")
                return redirect("voucher_edit", voucher_no=voucher.voucher_no)

            total_dr = sum(_q(e.debit) for e in entries)
            total_cr = sum(_q(e.credit) for e in entries)

            if total_dr != total_cr or total_dr <= ZERO:
                messages.error(request, "Journal entries must be balanced.")
                return redirect("voucher_edit", voucher_no=voucher.voucher_no)

            JournalEntry.objects.filter(journal=voucher).delete()
            for entry in entries:
                entry.journal = voucher
                entry.created_by = request.user
                entry.save()

            messages.success(request, f"Voucher {voucher.voucher_no} updated successfully.")
            return redirect("voucher_detail", voucher_no=voucher.voucher_no)
    else:
        form = VoucherCreateForm(instance=voucher)
        formset = JournalEntryFormSet(instance=voucher)

    return render(
        request,
        "finance/voucher_form.html",
        {"form": form, "formset": formset, "title": f"Edit Voucher {voucher.voucher_no}"},
    )


@user_passes_test(admin_required)
@transaction.atomic
def voucher_delete(request, voucher_no):
    voucher = get_object_or_404(Journal, Q(voucher_no=voucher_no) | Q(journal_id=voucher_no))

    if voucher.status in ["posted", "cancelled"]:
        messages.error(request, "Posted or cancelled voucher cannot be deleted.")
        return redirect("voucher_detail", voucher_no=voucher.voucher_no)

    if request.method == "POST":
        voucher.delete()
        messages.success(request, "Voucher deleted successfully.")
        return redirect("voucher_list")

    return render(request, "finance/voucher_delete.html", {"voucher": voucher})


@user_passes_test(admin_required)
@transaction.atomic
def voucher_post(request, voucher_no):
    voucher = get_object_or_404(Journal, Q(voucher_no=voucher_no) | Q(journal_id=voucher_no))
    try:
        voucher.post(user=request.user)
        messages.success(request, f"Voucher {voucher.voucher_no} posted successfully.")
    except Exception as e:
        messages.error(request, f"Posting failed: {e}")

    return redirect("voucher_detail", voucher_no=voucher.voucher_no)


@user_passes_test(admin_required)
@transaction.atomic
def voucher_cancel(request, voucher_no):
    voucher = get_object_or_404(Journal, Q(voucher_no=voucher_no) | Q(journal_id=voucher_no))

    if request.method == "POST":
        reason = request.POST.get("reason", "")
        voucher.status = "cancelled"
        voucher.cancelled_by = request.user
        voucher.cancelled_at = timezone.now()
        voucher.cancel_reason = reason
        voucher.save()

        messages.success(request, f"Voucher {voucher.voucher_no} cancelled.")
        return redirect("voucher_detail", voucher_no=voucher.voucher_no)

    return render(request, "finance/voucher_cancel.html", {"voucher": voucher})


# ==========================================================
# PRINT & PDF GENERATION
# ==========================================================

@user_passes_test(admin_required)
def voucher_print(request, voucher_no):
    voucher = get_object_or_404(Journal, Q(voucher_no=voucher_no) | Q(journal_id=voucher_no))
    return render(request, "finance/voucher_print.html", {"voucher": voucher, "journal": voucher})


@user_passes_test(admin_required)
def voucher_pdf(request, voucher_no):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    voucher = get_object_or_404(Journal, Q(voucher_no=voucher_no) | Q(journal_id=voucher_no))
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{voucher.voucher_no}.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 50

    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(50, y, "ACCOUNTING VOUCHER")

    y -= 40
    pdf.setFont("Helvetica", 11)
    pdf.drawString(50, y, f"Voucher No: {voucher.voucher_no}")
    y -= 20
    pdf.drawString(50, y, f"Type: {voucher.get_voucher_type_display()}")
    y -= 20
    pdf.drawString(50, y, f"Status: {voucher.status.upper()}")
    y -= 20
    pdf.drawString(50, y, f"Date: {voucher.created_at.strftime('%d-%m-%Y')}")

    y -= 40
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(50, y, "Account")
    pdf.drawString(300, y, "Debit")
    pdf.drawString(420, y, "Credit")

    y -= 20
    pdf.setFont("Helvetica", 10)
    for entry in voucher.entries.all():
        pdf.drawString(50, y, entry.account.name[:35])
        pdf.drawRightString(360, y, str(entry.debit))
        pdf.drawRightString(480, y, str(entry.credit))
        y -= 18
        if y < 80:
            pdf.showPage()
            y = height - 50

    y -= 20
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(50, y, "TOTAL")
    pdf.drawRightString(360, y, str(voucher.total_debit))
    pdf.drawRightString(480, y, str(voucher.total_credit))

    pdf.save()
    return response


def get_account_journals(request, account_id):
    journals = Journal.objects.filter(
        entries__account_id=account_id,
        status="posted"
    ).distinct().order_by("-created_at")

    data = [
        {
            "journal_id": j.journal_id,
            "voucher_no": j.voucher_no,
            "reference_no": j.reference_no,
            "date": j.created_at.strftime("%d-%m-%Y"),
        }
        for j in journals
    ]
    return JsonResponse(data, safe=False)





# ==========================================================
# 📊 EXPORT ALL LEDGERS TO EXCEL (ADMIN ONLY)
# ==========================================================

@login_required
@user_passes_test(admin_required)
def export_all_ledgers_excel(request):
    """
    Export all ledgers (Customer, Product, Cash, Supplier, Expense, etc.)
    into a single formatted Excel sheet with real-time calculations.
    """
    from_date_str = request.GET.get("from_date", "").strip()
    to_date_str = request.GET.get("to_date", "").strip()
    search_q = request.GET.get("search", "").strip()

    # Create Workbook & Sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "All General Ledgers"

    # Ensure Gridlines are visible
    ws.views.sheetView[0].showGridLines = True

    # ------------------ STYLES ------------------
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark Navy Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    acc_banner_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light Blue
    acc_banner_font = Font(name="Calibri", size=11, bold=True, color="000000")

    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    total_font = Font(name="Calibri", size=11, bold=True)

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    double_bottom_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    # ------------------ TITLE & INFO ------------------
    ws.merge_cells("A1:J1")
    ws["A1"] = "MASTER GENERAL LEDGER STATEMENT"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Sub-header Date Info
    ws.merge_cells("A2:J2")
    date_info = f"Generated On: {timezone.now().strftime('%d-%b-%Y %I:%M %p')}"
    if from_date_str or to_date_str:
        date_info += f" | Period: {from_date_str or 'Start'} to {to_date_str or 'Present'}"
    else:
        date_info += " | Period: All Time (Real-Time)"

    ws["A2"] = date_info
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ------------------ TABLE HEADERS ------------------
    headers = [
        "SL No",
        "Account Code",
        "Account / Product Name",
        "Date",
        "Journal ID",
        "Voucher No",
        "Particulars",
        "Debit (৳)",
        "Credit (৳)",
        "Balance (৳)"
    ]

    ws.append([])  # Empty Row 3
    ws.append(headers)  # Row 4
    ws.row_dimensions[4].height = 24

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # ------------------ FETCH DATA & POPULATE ------------------
    accounts_qs = Account.objects.all().order_by("type1", "account_code", "name")
    if search_q:
        accounts_qs = accounts_qs.filter(
            Q(account_code__icontains=search_q) | Q(name__icontains=search_q)
        )

    current_row = 5
    sl_no = 1
    grand_total_debit = ZERO
    grand_total_credit = ZERO

    for acc in accounts_qs:
        # Check if an Opening Balance Journal Voucher already exists
        has_op_journal = JournalEntry.objects.filter(
            account=acc,
            journal__status="posted",
            journal__notes__icontains="Opening Balance"
        ).exists()

        # 1. Base opening balance from Account model
        if has_op_journal:
            init_dr = ZERO
            init_cr = ZERO
        else:
            init_dr = _q(acc.opening_balance) if acc.opening_balance_type == "debit" else ZERO
            init_cr = _q(acc.opening_balance) if acc.opening_balance_type == "credit" else ZERO

        # 2. Prior entries if from_date filter is used
        entries_before_dr = ZERO
        entries_before_cr = ZERO
        if from_date_str:
            prior_entries = JournalEntry.objects.filter(
                account=acc,
                journal__status="posted",
                entry_date__date__lt=from_date_str
            ).aggregate(dr=Sum("debit"), cr=Sum("credit"))
            entries_before_dr = _q(prior_entries["dr"])
            entries_before_cr = _q(prior_entries["cr"])

        total_prev_dr = init_dr + entries_before_dr
        total_prev_cr = init_cr + entries_before_cr

        # Account Type-wise Starting/Opening Balance calculation
        if acc.type1 in ["asset", "expense"]:
            acc_running_balance = _q(total_prev_dr - total_prev_cr)
        else:
            acc_running_balance = _q(total_prev_cr - total_prev_dr)

        # 3. Filter entries within date range
        entry_filter = Q(account=acc, journal__status="posted")
        if from_date_str:
            entry_filter &= Q(entry_date__date__gte=from_date_str)
        if to_date_str:
            entry_filter &= Q(entry_date__date__lte=to_date_str)

        entries = JournalEntry.objects.filter(entry_filter).select_related("journal").order_by("entry_date", "id")

        # Skip account if no opening balance and no entries
        if acc_running_balance == ZERO and not entries.exists():
            continue

        # --- ACCOUNT BANNER / SECTION HEADER ---
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        banner_cell = ws.cell(row=current_row, column=1)
        banner_cell.value = f"▶ [{acc.account_code or 'N/A'}] {acc.name}  |  Type: {acc.get_type1_display()} ({acc.get_type2_display()})"
        banner_cell.font = acc_banner_font
        banner_cell.fill = acc_banner_fill
        banner_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 20

        for c in range(1, 11):
            ws.cell(row=current_row, column=c).border = thin_border

        current_row += 1

        # --- OPENING BALANCE ROW ---
        ws.append([
            sl_no,
            acc.account_code or "",
            acc.name,
            from_date_str or (acc.created_at.strftime("%d-%m-%Y") if acc.created_at else "Opening"),
            "-",
            "-",
            "Opening Balance",
            total_prev_dr if total_prev_dr > ZERO else "",
            total_prev_cr if total_prev_cr > ZERO else "",
            acc_running_balance
        ])

        for col_i in range(1, 11):
            c_cell = ws.cell(row=current_row, column=col_i)
            c_cell.border = thin_border
            c_cell.font = Font(name="Calibri", size=10, italic=True)
            if col_i in [8, 9, 10] and isinstance(c_cell.value, (int, float, Decimal)):
                c_cell.number_format = '#,##0.00'
                c_cell.alignment = Alignment(horizontal="right")
            elif col_i in [1, 2, 4, 5, 6]:
                c_cell.alignment = Alignment(horizontal="center")

        current_row += 1
        sl_no += 1

        # --- TRANSACTION ROWS ---
        for entry in entries:
            dr = _q(entry.debit)
            cr = _q(entry.credit)

            if acc.type1 in ["asset", "expense"]:
                acc_running_balance = _q(acc_running_balance + dr - cr)
            else:
                acc_running_balance = _q(acc_running_balance + cr - dr)

            grand_total_debit = _q(grand_total_debit + dr)
            grand_total_credit = _q(grand_total_credit + cr)

            date_val = entry.entry_date.strftime("%d-%m-%Y") if entry.entry_date else ""
            journal_id_val = entry.journal.journal_id if entry.journal else ""
            voucher_no_val = entry.voucher_no or (entry.journal.voucher_no if entry.journal else "") or "-"
            narration_val = entry.narration or (entry.journal.notes if entry.journal else "") or "-"

            ws.append([
                sl_no,
                acc.account_code or "",
                acc.name,
                date_val,
                journal_id_val,
                voucher_no_val,
                narration_val,
                dr if dr > ZERO else "",
                cr if cr > ZERO else "",
                acc_running_balance
            ])

            for col_i in range(1, 11):
                c_cell = ws.cell(row=current_row, column=col_i)
                c_cell.border = thin_border
                c_cell.font = Font(name="Calibri", size=10)
                if col_i in [8, 9, 10] and isinstance(c_cell.value, (int, float, Decimal)):
                    c_cell.number_format = '#,##0.00'
                    c_cell.alignment = Alignment(horizontal="right")
                elif col_i in [1, 2, 4, 5, 6]:
                    c_cell.alignment = Alignment(horizontal="center")

            current_row += 1
            sl_no += 1

        # Small spacing after each account
        current_row += 1
        ws.row_dimensions[current_row - 1].height = 8

    # ------------------ GRAND TOTAL ROW ------------------
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    tot_label = ws.cell(row=current_row, column=1)
    tot_label.value = "GRAND TOTAL (ALL TRANSACTIONS)"
    tot_label.font = total_font
    tot_label.alignment = Alignment(horizontal="right", vertical="center")

    dr_cell = ws.cell(row=current_row, column=8, value=grand_total_debit)
    dr_cell.font = total_font
    dr_cell.number_format = '#,##0.00'
    dr_cell.alignment = Alignment(horizontal="right")

    cr_cell = ws.cell(row=current_row, column=9, value=grand_total_credit)
    cr_cell.font = total_font
    cr_cell.number_format = '#,##0.00'
    cr_cell.alignment = Alignment(horizontal="right")

    bal_cell = ws.cell(row=current_row, column=10, value=_q(grand_total_debit - grand_total_credit))
    bal_cell.font = total_font
    bal_cell.number_format = '#,##0.00'
    bal_cell.alignment = Alignment(horizontal="right")

    for col_i in range(1, 11):
        cell = ws.cell(row=current_row, column=col_i)
        cell.fill = total_fill
        cell.border = double_bottom_border

    ws.row_dimensions[current_row].height = 24

    # ------------------ AUTO COLUMN WIDTHS ------------------
    col_widths = {
        "A": 8,  # SL No
        "B": 15,  # Account Code
        "C": 30,  # Account / Product Name
        "D": 14,  # Date
        "E": 15,  # Journal ID
        "F": 18,  # Voucher No
        "G": 38,  # Particulars
        "H": 16,  # Debit
        "I": 16,  # Credit
        "J": 18,  # Balance
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ------------------ HTTP RESPONSE ------------------
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    response["Content-Disposition"] = f'attachment; filename="All_Ledgers_Report_{timestamp}.xlsx"'
    wb.save(response)
    return response

